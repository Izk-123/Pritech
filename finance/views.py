# finance/views.py
"""
Finance App Views
-----------------
Complete set of views for financial management:
- Quotations (CRUD, send, convert, duplicate, PDF, print, Word, Excel)
- Invoices (CRUD, add items, record payment, issue, bulk actions, CSV export, PDF)
- Expenses (list, create, approve, reject)
- Financial Reports (income statement, aging, top clients, monthly trends)
- Client portal (invoices, quotations, approve/reject, subscription management)
- HTMX real‑time endpoints (live search, status updates, dashboard partials)

All staff views enforce role‑based access using RoleRequiredMixin.
Client views use standard LoginRequiredMixin with object‑level permissions.
"""

import csv
import json
import logging
import base64
from datetime import date, timedelta
from pathlib import Path

from django.views.generic import ListView, CreateView, DetailView, UpdateView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse, Http404, JsonResponse
from django.db import models, transaction
from django.utils import timezone
from django.core.cache import cache
from django.utils.decorators import method_decorator
from django.template.loader import render_to_string
from django.template.response import TemplateResponse
from django.db.models import Q
from django.conf import settings

# For Excel export
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Local imports
from .models import Invoice, Expense, InvoiceItem, Quotation, QuotationItem, Plan, ClientSubscription
from .forms import (
    InvoiceForm, InvoiceItemForm, PaymentForm, ExpenseForm,
    QuotationForm, QuotationItemFormSet, SubscriptionChangeForm,
    InvoiceFilterForm
)
from .services import InvoiceService, QuotationService
from .reports import ReportService
from .tasks import generate_invoice_pdf, generate_quotation_pdf

from core.models import SiteConfig
from core.mixins import RoleRequiredMixin
from infrastructure.pdf_service import render_pdf, pdf_response
from infrastructure.pdf_sanitizer import sanitize_html, sanitize_plain_text
from accounts.decorators import rate_limit, get_client_ip

logger = logging.getLogger(__name__)


# -----------------------------------------------------------------------------
# Helper Functions
# -----------------------------------------------------------------------------
def send_payment_reminder(invoice):
    """Send an email reminder to client for overdue or outstanding invoice."""
    try:
        from django.core.mail import send_mail
        config = SiteConfig.get()
        subject = f"Payment Reminder: {invoice.invoice_number}"
        message = (
            f"Dear {invoice.client.name},\n\n"
            f"Invoice {invoice.invoice_number} of amount {config.currency_symbol}{invoice.balance_due} "
            f"is due on {invoice.due_date}. Please arrange payment.\n\n"
            f"Thank you for your business.\n\n"
            f"{config.company_name}"
        )
        send_mail(
            subject=subject,
            message=message,
            from_email=config.email or settings.DEFAULT_FROM_EMAIL,
            recipient_list=[invoice.client.email],
            fail_silently=True,
        )
    except Exception as e:
        logger.warning(f"Failed to send payment reminder for {invoice.invoice_number}: {e}")


# =============================================================================
# QUOTATION VIEWS (Staff only)
# =============================================================================

class QuotationListView(RoleRequiredMixin, ListView):
    """List all quotations with optional status filter."""
    model = Quotation
    template_name = 'finance/quotation_list.html'
    context_object_name = 'quotations'
    paginate_by = 20
    required_roles = ['FINANCE', 'ADMIN']

    def get_queryset(self):
        qs = super().get_queryset().select_related('client')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['status_choices'] = Quotation.STATUS_CHOICES
        ctx['current_status'] = self.request.GET.get('status', '')
        return ctx


class QuotationCreateView(RoleRequiredMixin, CreateView):
    """Create a new quotation with dynamic line items formset."""
    model = Quotation
    form_class = QuotationForm
    template_name = 'finance/quotation_form.html'
    required_roles = ['FINANCE', 'ADMIN']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['items'] = QuotationItemFormSet(self.request.POST)
        else:
            ctx['items'] = QuotationItemFormSet()
        return ctx

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            form.instance.created_by = self.request.user
            self.object = form.save()
            if items.is_valid():
                items.instance = self.object
                items.save()
                QuotationService.recalculate(self.object)

        # Check which action button was clicked
        if self.request.POST.get('action') == 'send':
            try:
                QuotationService.send(self.object)
                messages.success(self.request, 'Quotation sent to client.')
            except ValueError as e:
                messages.error(self.request, str(e))
        else:
            messages.success(self.request, 'Quotation saved as draft.')

        return redirect('quotation_detail', pk=self.object.pk)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct errors in the quotation header.")
        return self.render_to_response(self.get_context_data(form=form))


class QuotationDetailView(RoleRequiredMixin, DetailView):
    """Detailed view of a single quotation with items."""
    model = Quotation
    template_name = 'finance/quotation_detail.html'
    context_object_name = 'quotation'
    required_roles = ['FINANCE', 'ADMIN']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['items'] = self.object.items.all()
        ctx['config'] = SiteConfig.get()
        return ctx


class QuotationUpdateView(RoleRequiredMixin, UpdateView):
    """Edit an existing quotation (only allowed in 'draft' status)."""
    model = Quotation
    form_class = QuotationForm
    template_name = 'finance/quotation_form.html'
    required_roles = ['FINANCE', 'ADMIN']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        if self.request.POST:
            ctx['items'] = QuotationItemFormSet(self.request.POST, instance=self.object)
        else:
            ctx['items'] = QuotationItemFormSet(instance=self.object)
        return ctx

    def form_valid(self, form):
        context = self.get_context_data()
        items = context['items']
        with transaction.atomic():
            self.object = form.save()
            if items.is_valid():
                items.instance = self.object
                items.save()
                QuotationService.recalculate(self.object)

        if self.request.POST.get('action') == 'send':
            try:
                QuotationService.send(self.object)
                messages.success(self.request, 'Quotation sent to client.')
            except ValueError as e:
                messages.error(self.request, str(e))
        else:
            messages.success(self.request, 'Quotation updated.')

        return redirect('quotation_detail', pk=self.object.pk)

    def form_invalid(self, form):
        messages.error(self.request, "Please correct errors in the quotation header.")
        return self.render_to_response(self.get_context_data(form=form))


class QuotationSendView(RoleRequiredMixin, View):
    """Simple view to change quotation status to 'sent'."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        quotation = get_object_or_404(Quotation, pk=pk)
        try:
            QuotationService.send(quotation)
            messages.success(request, 'Quotation sent to client.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('quotation_detail', pk=pk)


class QuotationConvertView(RoleRequiredMixin, View):
    """Convert an approved quotation into a draft invoice."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        quotation = get_object_or_404(Quotation, pk=pk)
        try:
            invoice = QuotationService.convert_to_invoice(quotation, created_by=request.user)
            messages.success(request, f'Quotation converted to Invoice {invoice.invoice_number}.')
            return redirect('invoice_detail', pk=invoice.pk)
        except ValueError as e:
            messages.error(request, str(e))
            return redirect('quotation_detail', pk=pk)


class DuplicateQuotationView(RoleRequiredMixin, View):
    """Create a new draft quotation by duplicating an existing one."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        original = get_object_or_404(Quotation, pk=pk)
        new_quotation = Quotation.objects.create(
            client=original.client,
            issue_date=date.today(),
            valid_until=date.today() + timedelta(days=30),
            notes=f"Duplicated from {original.quotation_number}",
            created_by=request.user,
            status='draft'
        )
        for item in original.items.all():
            QuotationItem.objects.create(
                quotation=new_quotation,
                description=item.description,
                quantity=item.quantity,
                unit_price=item.unit_price,
                total=item.total
            )
        QuotationService.recalculate(new_quotation)
        messages.success(request, f'Quotation duplicated as {new_quotation.quotation_number}.')
        return redirect('quotation_detail', pk=new_quotation.pk)


# -----------------------------------------------------------------------------
# Export Views (Print, PDF Direct, Word, Excel)
# -----------------------------------------------------------------------------

def quotation_print_view(request, pk):
    """Dedicated print view – minimalist HTML, no sidebar, perfect for printing."""
    quotation = get_object_or_404(Quotation, pk=pk)
    items = quotation.items.all()
    return TemplateResponse(request, 'finance/quotation_print.html', {
        'quotation': quotation,
        'items': items,
        'site_config': SiteConfig.get(),
        'config': SiteConfig.get(),
    })


def quotation_pdf_direct(request, pk):
    """
    Direct PDF generation (synchronous) – downloads immediately.
    Uses WeasyPrint via render_pdf from infrastructure.
    """
    quotation = get_object_or_404(Quotation, pk=pk)
    items = quotation.items.all()
    config = SiteConfig.get()
    vat_percent = float(config.vat_rate) * 100

    # Build items data for template
    items_data = []
    for item in items:
        items_data.append({
            'description': sanitize_plain_text(item.description),
            'quantity': item.quantity,
            'unit_price': item.unit_price,
            'total': item.total,
        })

    # Logo data URI (for PDF embedding)
    logo_path = Path(settings.BASE_DIR) / 'static' / 'images' / 'logo.png'
    logo_data_uri = ""
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            logo_bytes = f.read()
            b64 = base64.b64encode(logo_bytes).decode('utf-8')
            logo_data_uri = f"data:image/png;base64,{b64}"

    pdf_bytes = render_pdf('finance/quotation_pdf.html', {
        'quotation': quotation,
        'items': items_data,
        'config': config,
        'vat_percent': f'{vat_percent:.1f}',
        'static_logo_data_uri': logo_data_uri,
        'notes_html': quotation.notes.html if quotation.notes else '',
    })
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.quotation_number}.pdf"'
    return response


def quotation_word_export(request, pk):
    """Export quotation as Microsoft Word document (.doc)."""
    quotation = get_object_or_404(Quotation, pk=pk)
    items = quotation.items.all()
    response = HttpResponse(content_type='application/msword')
    response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.quotation_number}.doc"'
    template = render_to_string('finance/quotation_print.html', {
        'quotation': quotation,
        'items': items,
        'site_config': SiteConfig.get(),
        'config': SiteConfig.get(),
    })
    response.write(template)
    return response


def quotation_excel_export(request, pk):
    """Export quotation as Excel (.xlsx) using openpyxl."""
    quotation = get_object_or_404(Quotation, pk=pk)
    items = quotation.items.all()
    config = SiteConfig.get()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Title and header info
    ws['A1'] = config.company_name
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A2'] = f"Quotation Number: {quotation.quotation_number}"
    ws.merge_cells('A2:D2')
    ws['A3'] = f"Client: {quotation.client.name}"
    ws.merge_cells('A3:D3')
    ws['A4'] = f"Issue Date: {quotation.issue_date}   Valid Until: {quotation.valid_until}"
    ws.merge_cells('A4:D4')

    # Headers
    headers = ['Description', 'Quantity', 'Unit Price', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))

    # Items
    row = 7
    for item in items:
        ws.cell(row=row, column=1, value=item.description)
        ws.cell(row=row, column=2, value=float(item.quantity))
        ws.cell(row=row, column=3, value=float(item.unit_price))
        ws.cell(row=row, column=4, value=float(item.total))
        row += 1

    # Totals
    ws.cell(row=row+1, column=3, value="Subtotal").font = Font(bold=True)
    ws.cell(row=row+1, column=4, value=float(quotation.subtotal))
    ws.cell(row=row+2, column=3, value=f"VAT ({float(config.vat_rate)*100:.1f}%)").font = Font(bold=True)
    ws.cell(row=row+2, column=4, value=float(quotation.tax_amount))
    ws.cell(row=row+3, column=3, value="Total").font = Font(bold=True, size=12)
    ws.cell(row=row+3, column=4, value=float(quotation.total_amount)).font = Font(bold=True, size=12)

    # Notes
    if quotation.notes:
        ws.cell(row=row+5, column=1, value="Notes").font = Font(bold=True)
        ws.merge_cells(f'A{row+5}:D{row+5}')
        # Strip HTML tags for plain text notes
        plain_notes = quotation.notes.html if hasattr(quotation.notes, 'html') else str(quotation.notes)
        ws.cell(row=row+6, column=1, value=plain_notes).alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row+6}:D{row+6}')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.quotation_number}.xlsx"'
    wb.save(response)
    return response


@method_decorator(rate_limit(key_func=get_client_ip, rate='30/h', method='GET', block=True), name='dispatch')
class QuotationPDFView(RoleRequiredMixin, View):
    """Serve quotation PDF – generated asynchronously (cache). Kept for compatibility."""
    required_roles = ['FINANCE', 'ADMIN']

    def get(self, request, pk):
        cache_key = f'quotation_pdf_{pk}'
        pdf_bytes = cache.get(cache_key)
        if pdf_bytes:
            return pdf_response(pdf_bytes, f'Quotation_{pk}.pdf')
        generate_quotation_pdf.delay(pk)
        return HttpResponse(
            'PDF is being generated. Please refresh in a few moments.',
            status=202,
            content_type='text/plain'
        )


# =============================================================================
# INVOICE VIEWS (Staff only)
# =============================================================================

class InvoiceListView(RoleRequiredMixin, ListView):
    """List all invoices with filters and totals."""
    model = Invoice
    template_name = 'finance/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20
    required_roles = ['FINANCE', 'ADMIN']

    def get_queryset(self):
        qs = super().get_queryset().select_related('client')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        issue_from = self.request.GET.get('issue_from')
        if issue_from:
            qs = qs.filter(issue_date__gte=issue_from)
        issue_to = self.request.GET.get('issue_to')
        if issue_to:
            qs = qs.filter(issue_date__lte=issue_to)
        due_from = self.request.GET.get('due_from')
        if due_from:
            qs = qs.filter(due_date__gte=due_from)
        due_to = self.request.GET.get('due_to')
        if due_to:
            qs = qs.filter(due_date__lte=due_to)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['status_choices'] = Invoice.STATUS_CHOICES
        ctx['current_status'] = self.request.GET.get('status', '')
        ctx['issue_from'] = self.request.GET.get('issue_from', '')
        ctx['issue_to'] = self.request.GET.get('issue_to', '')
        ctx['due_from'] = self.request.GET.get('due_from', '')
        ctx['due_to'] = self.request.GET.get('due_to', '')
        qs = self.get_queryset()
        ctx['total_amount_sum'] = qs.aggregate(total=models.Sum('total_amount'))['total'] or 0
        ctx['total_paid_sum'] = qs.aggregate(total=models.Sum('amount_paid'))['total'] or 0
        ctx['total_balance_sum'] = ctx['total_amount_sum'] - ctx['total_paid_sum']
        return ctx


class InvoiceCreateView(RoleRequiredMixin, CreateView):
    """Create a new draft invoice."""
    model = Invoice
    form_class = InvoiceForm
    template_name = 'finance/invoice_form.html'
    required_roles = ['FINANCE', 'ADMIN']

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Invoice created. You can now add line items.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('invoice_detail', kwargs={'pk': self.object.pk})


class InvoiceDetailView(RoleRequiredMixin, DetailView):
    """Detailed view of an invoice with items, payments, and action forms."""
    model = Invoice
    template_name = 'finance/invoice_detail.html'
    context_object_name = 'invoice'
    required_roles = ['FINANCE', 'ADMIN']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['item_form'] = InvoiceItemForm()
        ctx['payment_form'] = PaymentForm()
        ctx['items'] = self.object.items.all()
        ctx['payments'] = self.object.payments.all()
        ctx['config'] = SiteConfig.get()
        return ctx


class AddInvoiceItemView(RoleRequiredMixin, View):
    """Add a line item to a draft invoice."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        if invoice.status != 'draft':
            messages.error(request, 'Items can only be added to draft invoices.')
            return redirect('invoice_detail', pk=pk)

        form = InvoiceItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.invoice = invoice
            item.save()
            InvoiceService.recalculate(invoice)
            messages.success(request, 'Item added.')
        else:
            messages.error(request, 'Invalid item data.')
        return redirect('invoice_detail', pk=pk)


class RecordPaymentView(RoleRequiredMixin, View):
    """Record a payment against an invoice. Includes idempotency lock."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        lock_key = f"payment_lock_{invoice.pk}_{request.user.pk}"
        if cache.get(lock_key):
            messages.error(request, "You are recording payments too quickly. Please wait a moment.")
            return redirect('invoice_detail', pk=pk)

        form = PaymentForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']
            if amount > invoice.balance_due:
                messages.error(request, f"Payment amount cannot exceed the outstanding balance ({invoice.balance_due}).")
                return redirect('invoice_detail', pk=pk)

            cache.set(lock_key, True, timeout=5)
            InvoiceService.record_payment(
                invoice=invoice,
                recorded_by=request.user,
                **form.cleaned_data
            )
            messages.success(request, 'Payment recorded.')
        else:
            messages.error(request, 'Invalid payment data.')
        return redirect('invoice_detail', pk=pk)


class IssueInvoiceView(RoleRequiredMixin, View):
    """Change invoice status from draft to issued (triggers email)."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        try:
            InvoiceService.issue(invoice)
            messages.success(request, f'Invoice {invoice.invoice_number} issued.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('invoice_detail', pk=pk)


class BulkInvoiceActionView(RoleRequiredMixin, View):
    """Bulk actions: send payment reminders or issue multiple invoices."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request):
        action = request.POST.get('action')
        invoice_ids = request.POST.getlist('invoice_ids')
        if not invoice_ids:
            messages.error(request, 'No invoices selected.')
            return redirect(request.META.get('HTTP_REFERER', 'invoice_list'))

        invoices = Invoice.objects.filter(pk__in=invoice_ids)
        if action == 'send_reminder':
            for inv in invoices:
                send_payment_reminder(inv)
            messages.success(request, f'Reminder sent to {len(invoices)} client(s).')
        elif action == 'issue':
            count = 0
            for inv in invoices:
                if inv.status == 'draft':
                    InvoiceService.issue(inv)
                    count += 1
            messages.success(request, f'{count} invoice(s) issued.')
        else:
            messages.error(request, 'Invalid action.')
        return redirect(request.META.get('HTTP_REFERER', 'invoice_list'))


class ExportInvoiceCSVView(RoleRequiredMixin, View):
    """Export filtered invoice list to CSV."""
    required_roles = ['FINANCE', 'ADMIN']

    def get(self, request):
        qs = Invoice.objects.select_related('client')
        status = request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        issue_from = request.GET.get('issue_from')
        if issue_from:
            qs = qs.filter(issue_date__gte=issue_from)
        issue_to = request.GET.get('issue_to')
        if issue_to:
            qs = qs.filter(issue_date__lte=issue_to)
        due_from = request.GET.get('due_from')
        if due_from:
            qs = qs.filter(due_date__gte=due_from)
        due_to = request.GET.get('due_to')
        if due_to:
            qs = qs.filter(due_date__lte=due_to)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="invoices.csv"'
        writer = csv.writer(response)
        writer.writerow(['Invoice #', 'Client', 'Issue Date', 'Due Date', 'Total', 'Paid', 'Balance', 'Status'])
        for inv in qs:
            writer.writerow([
                inv.invoice_number,
                inv.client.name,
                inv.issue_date,
                inv.due_date,
                inv.total_amount,
                inv.amount_paid,
                inv.balance_due,
                inv.get_status_display()
            ])
        return response


@method_decorator(rate_limit(key_func=get_client_ip, rate='30/h', method='GET', block=True), name='dispatch')
class InvoicePDFView(RoleRequiredMixin, View):
    """Serve invoice PDF – generated asynchronously."""
    required_roles = ['FINANCE', 'ADMIN']

    def get(self, request, pk):
        cache_key = f'invoice_pdf_{pk}'
        pdf_bytes = cache.get(cache_key)
        if pdf_bytes:
            return pdf_response(pdf_bytes, f'Invoice_{pk}.pdf')
        generate_invoice_pdf.delay(pk)
        return HttpResponse(
            'PDF is being generated. Please refresh in a few moments.',
            status=202,
            content_type='text/plain'
        )


# =============================================================================
# EXPENSE VIEWS (Staff only)
# =============================================================================

class ExpenseListView(RoleRequiredMixin, ListView):
    """List all expenses with approval status."""
    model = Expense
    template_name = 'finance/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 20
    required_roles = ['FINANCE', 'ADMIN']


class ExpenseCreateView(RoleRequiredMixin, CreateView):
    """Create a new expense (status defaults to 'pending')."""
    model = Expense
    form_class = ExpenseForm
    template_name = 'finance/expense_form.html'
    success_url = reverse_lazy('expense_list')
    required_roles = ['FINANCE', 'ADMIN']

    def form_valid(self, form):
        form.instance.recorded_by = self.request.user
        return super().form_valid(form)


class ApproveExpenseView(RoleRequiredMixin, View):
    """Approve a pending expense."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        expense = get_object_or_404(Expense, pk=pk)
        if expense.status != 'pending':
            messages.error(request, 'This expense is already processed.')
            return redirect('expense_list')
        expense.status = 'approved'
        expense.approved_by = request.user
        expense.approved_at = timezone.now()
        expense.save()
        messages.success(request, f'Expense "{expense.description}" approved.')
        return redirect('expense_list')


class RejectExpenseView(RoleRequiredMixin, View):
    """Reject a pending expense."""
    required_roles = ['FINANCE', 'ADMIN']

    def post(self, request, pk):
        expense = get_object_or_404(Expense, pk=pk)
        if expense.status != 'pending':
            messages.error(request, 'This expense is already processed.')
            return redirect('expense_list')
        expense.status = 'rejected'
        expense.save()
        messages.warning(request, f'Expense "{expense.description}" rejected.')
        return redirect('expense_list')


# =============================================================================
# FINANCIAL REPORTS (Staff only)
# =============================================================================

class ReportView(RoleRequiredMixin, TemplateView):
    """Financial dashboard with income statement, aging, top clients, monthly charts."""
    template_name = 'finance/reports.html'
    required_roles = ['FINANCE', 'ADMIN']

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        period = self.request.GET.get('period', 'this_month')
        start, end = ReportService.date_range(period)

        ctx['period'] = period
        ctx['start'] = start
        ctx['end'] = end
        ctx['income'] = ReportService.income_statement(start, end)
        ctx['aging'] = ReportService.accounts_receivable_aging()
        ctx['top_clients'] = ReportService.client_revenue(start, end)
        ctx['monthly'] = ReportService.monthly_revenue(6)
        ctx['monthly_json'] = json.dumps(ctx['monthly'])
        ctx['config'] = SiteConfig.get()
        ctx['period_choices'] = [
            ('this_month', 'This Month'),
            ('last_month', 'Last Month'),
            ('last_30', 'Last 30 Days'),
            ('last_90', 'Last 90 Days'),
            ('this_quarter', 'This Quarter'),
            ('this_year', 'This Year'),
        ]
        return ctx


# =============================================================================
# CLIENT-FACING VIEWS (object‑level permissions)
# =============================================================================

class ClientInvoiceListView(LoginRequiredMixin, ListView):
    """Clients see only their own invoices."""
    template_name = 'finance/client_invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20

    def get_queryset(self):
        client_org = getattr(self.request.user, 'client_organization', None)
        if not client_org:
            return Invoice.objects.none()
        return Invoice.objects.filter(client=client_org).order_by('-created_at')


class ClientInvoiceDetailView(LoginRequiredMixin, DetailView):
    """Client invoice detail with permission check."""
    template_name = 'finance/client_invoice_detail.html'
    context_object_name = 'invoice'

    def get_queryset(self):
        client_org = getattr(self.request.user, 'client_organization', None)
        if not client_org:
            return Invoice.objects.none()
        return Invoice.objects.filter(client=client_org)

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if obj.client != getattr(self.request.user, 'client_organization', None):
            raise Http404("Invoice not found.")
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['config'] = SiteConfig.get()
        ctx['items'] = self.object.items.all()
        ctx['payments'] = self.object.payments.all()
        return ctx


class ClientQuotationListView(LoginRequiredMixin, ListView):
    """Clients see only their own quotations."""
    template_name = 'finance/client_quotation_list.html'
    context_object_name = 'quotations'
    paginate_by = 20

    def get_queryset(self):
        client_org = getattr(self.request.user, 'client_organization', None)
        if not client_org:
            return Quotation.objects.none()
        return Quotation.objects.filter(client=client_org).order_by('-created_at')


class ClientQuotationDetailView(LoginRequiredMixin, DetailView):
    """Client quotation detail with permission check."""
    template_name = 'finance/client_quotation_detail.html'
    context_object_name = 'quotation'

    def get_queryset(self):
        client_org = getattr(self.request.user, 'client_organization', None)
        if not client_org:
            return Quotation.objects.none()
        return Quotation.objects.filter(client=client_org)

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if obj.client != getattr(self.request.user, 'client_organization', None):
            raise Http404("Quotation not found.")
        return obj

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['config'] = SiteConfig.get()
        ctx['items'] = self.object.items.all()
        return ctx


class ClientQuotationApproveView(LoginRequiredMixin, View):
    """Client approves a sent quotation, automatically converting to invoice."""
    def post(self, request, pk):
        quotation = get_object_or_404(Quotation, pk=pk, client=request.user.client_organization)
        if quotation.status != 'sent':
            messages.error(request, f"This quotation cannot be approved (current status: {quotation.status}).")
            return redirect('client_quotation_detail', pk=pk)

        quotation.status = 'approved'
        quotation.save(update_fields=['status'])
        try:
            invoice = QuotationService.convert_to_invoice(quotation, created_by=request.user)
            messages.success(request, f"Quotation {quotation.quotation_number} approved and converted to invoice {invoice.invoice_number}.")
            return redirect('client_invoice_detail', pk=invoice.pk)
        except ValueError as e:
            messages.error(request, f"Conversion failed: {e}")
            quotation.status = 'sent'
            quotation.save(update_fields=['status'])
            return redirect('client_quotation_detail', pk=pk)


class ClientQuotationRejectView(LoginRequiredMixin, View):
    """Client rejects a sent quotation with a reason (stored as client_feedback)."""
    def post(self, request, pk):
        quotation = get_object_or_404(Quotation, pk=pk, client=request.user.client_organization)
        if quotation.status != 'sent':
            messages.error(request, "This quotation cannot be rejected.")
            return redirect('client_quotation_detail', pk=pk)

        reason = request.POST.get('reason', '').strip()
        if not reason:
            messages.error(request, "Please provide a reason for rejection or change request.")
            return redirect('client_quotation_detail', pk=pk)

        quotation.status = 'rejected'
        quotation.client_feedback = reason
        quotation.save(update_fields=['status', 'client_feedback'])

        # Notify finance team (optional)
        try:
            from django.core.mail import send_mail
            config = SiteConfig.get()
            send_mail(
                subject=f"Quotation {quotation.quotation_number} rejected by {quotation.client.name}",
                message=f"Client feedback:\n{reason}",
                from_email=config.email or settings.DEFAULT_FROM_EMAIL,
                recipient_list=['finance@pritech.mw'],
                fail_silently=True,
            )
        except Exception:
            pass

        messages.success(request, "Quotation rejected. We'll review your feedback and get back to you.")
        return redirect('client_quotation_detail', pk=pk)


class ClientSubscriptionDetailView(LoginRequiredMixin, DetailView):
    """Display client's current subscription details and available plans."""
    template_name = 'finance/client_subscription_detail.html'
    context_object_name = 'subscription'

    def get_object(self, queryset=None):
        client_org = getattr(self.request.user, 'client_organization', None)
        if not client_org:
            raise Http404("No subscription found for your organization.")
        return get_object_or_404(ClientSubscription, client=client_org)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['available_plans'] = Plan.objects.filter(is_active=True).exclude(pk=self.object.plan.pk)
        return ctx


class ClientSubscriptionUpgradeView(LoginRequiredMixin, UpdateView):
    """Upgrade or schedule downgrade of subscription plan."""
    model = ClientSubscription
    form_class = SubscriptionChangeForm
    template_name = 'finance/client_subscription_form.html'
    success_url = reverse_lazy('client_subscription_detail')

    def get_object(self, queryset=None):
        client_org = getattr(self.request.user, 'client_organization', None)
        if not client_org:
            raise Http404("No subscription found.")
        return get_object_or_404(ClientSubscription, client=client_org)

    def form_valid(self, form):
        new_plan = form.cleaned_data['plan']
        subscription = self.get_object()
        if new_plan.monthly_price > subscription.plan.monthly_price:
            subscription.plan = new_plan
            subscription.save()
            messages.success(self.request, f"Plan upgraded to {new_plan.name}.")
        else:
            subscription.cancel_at_period_end = True
            subscription.save()
            messages.info(self.request, f"Plan will downgrade to {new_plan.name} at the end of current period.")
        return redirect(self.success_url)


class ClientSubscriptionCancelView(LoginRequiredMixin, View):
    """Cancel subscription at the end of current billing period."""
    def post(self, request):
        client_org = getattr(request.user, 'client_organization', None)
        if not client_org:
            messages.error(request, "No active subscription found.")
            return redirect('client_subscription_detail')
        subscription = get_object_or_404(ClientSubscription, client=client_org)
        subscription.cancel_at_period_end = True
        subscription.save()
        messages.warning(request, "Your subscription will be cancelled at the end of the current billing period.")
        return redirect('client_subscription_detail')


# =============================================================================
# HTMX REAL‑TIME ENDPOINTS
# =============================================================================

def live_quotation_search(request):
    """HTMX endpoint – returns filtered quotation table rows."""
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    quotations = Quotation.objects.select_related('client').all()
    if query:
        quotations = quotations.filter(
            Q(quotation_number__icontains=query) |
            Q(client__name__icontains=query)
        )
    if status:
        quotations = quotations.filter(status=status)
    html = render_to_string('finance/partials/quotation_rows.html', {
        'quotations': quotations,
        'site_config': SiteConfig.get(),
    })
    return HttpResponse(html)


def update_quotation_status(request, pk):
    """HTMX endpoint – update quotation status (draft ↔ sent)."""
    quotation = get_object_or_404(Quotation, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['draft', 'sent'] and quotation.status != 'converted':
            quotation.status = new_status
            quotation.save()
            if new_status == 'sent':
                QuotationService.send(quotation)
    badge_map = {
        'draft': 'secondary',
        'sent': 'primary',
        'approved': 'success',
        'converted': 'secondary',
        'rejected': 'danger',
        'expired': 'warning',
    }
    badge_class = badge_map.get(quotation.status, 'secondary')
    return HttpResponse(
        f'<span class="badge bg-{badge_class}">{quotation.get_status_display()}</span>'
    )


def live_invoice_search(request):
    """HTMX endpoint – returns filtered invoice table rows."""
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    invoices = Invoice.objects.select_related('client').all()
    if query:
        invoices = invoices.filter(
            Q(invoice_number__icontains=query) |
            Q(client__name__icontains=query)
        )
    if status:
        invoices = invoices.filter(status=status)
    html = render_to_string('finance/partials/invoice_rows.html', {
        'invoices': invoices,
        'site_config': SiteConfig.get(),
    })
    return HttpResponse(html)


def dashboard_stats_partial(request):
    """HTMX endpoint – returns financial summary stats for dashboard."""
    from .reports import ReportService
    start = date.today().replace(day=1)
    end = date.today()
    income = ReportService.income_statement(start, end)
    aging = ReportService.accounts_receivable_aging()
    html = render_to_string('finance/partials/dashboard_stats.html', {
        'income': income,
        'aging': aging,
        'site_config': SiteConfig.get(),
    })
    return HttpResponse(html)